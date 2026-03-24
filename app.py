import streamlit as st
import pandas as pd
import json
import io
import os
from openpyxl import Workbook

st.set_page_config(page_title="Retail Sync", layout="wide")
st.title("🛒 Retail Sync")

RUTA_MASTER = "data/master.parquet"
RUTA_BD = "data/BD_ACTUALIZACION.parquet"


def crear_master_vacio():
    columnas = [
        "PK_ARTICULO", "ACTIVIDAD_COMERCIAL", "FAMILIA", "CATEGORIA", "SUBCATEGORIA",
        "NO_ARTI", "DESCRIPCION", "TIPO_CLASIF",
        "COMPRA_Q_2024", "COMPRA_Q_2025", "COMPRA_Q_2026",
        "VTA_YTD_2024", "VTA_YTD_2025", "VTA_YTD_2026",
        "VTA_Q_YTD_2024", "VTA_Q_YTD_2025", "VTA_Q_YTD_2026",
        "INVENTARIO_Q",
        "MUNDO_AC", "PRECIO_PROMOCIONAL", "DESCUENTO", "PORC_AHORRO",
        "FECHA_INICIO", "FECHA_FIN",
        "ACCION", "COMENTARIO"
    ]

    df = pd.DataFrame(columns=columnas)
    df.to_parquet(RUTA_MASTER, index=False)


if not os.path.exists(RUTA_MASTER):
    crear_master_vacio()
else:
    try:
        test = pd.read_parquet(RUTA_MASTER)
        if "ACTIVIDAD_COMERCIAL" not in test.columns:
            crear_master_vacio()
    except Exception:
        crear_master_vacio()


# ======================================================
# ACTUALIZAR MASTER DESDE BD
# ======================================================
def actualizar_master():
    master = pd.read_parquet(RUTA_MASTER)

    if master.empty:
        return

    if not os.path.exists(RUTA_BD):
        return

    bd = pd.read_parquet(RUTA_BD)

    columnas_update = [
        "FAMILIA", "CATEGORIA", "SUBCATEGORIA", "NO_ARTI", "DESCRIPCION", "TIPO_CLASIF",
        "COMPRA_Q_2024", "COMPRA_Q_2025", "COMPRA_Q_2026",
        "VTA_YTD_2024", "VTA_YTD_2025", "VTA_YTD_2026",
        "VTA_Q_YTD_2024", "VTA_Q_YTD_2025", "VTA_Q_YTD_2026",
        "INVENTARIO_Q"
    ]

    columnas_existentes = ["PK_ARTICULO"] + [c for c in columnas_update if c in bd.columns]

    if len(columnas_existentes) == 1:
        return

    master = master.merge(
        bd[columnas_existentes],
        on="PK_ARTICULO",
        how="left",
        suffixes=("", "_NEW")
    )

    for col in columnas_update:
        if f"{col}_NEW" in master.columns:
            if col in master.columns:
                master[col] = master[f"{col}_NEW"].combine_first(master[col])
            else:
                master[col] = master[f"{col}_NEW"]

    master = master.drop(columns=[c for c in master.columns if c.endswith("_NEW")])

    master.to_parquet(RUTA_MASTER, index=False)


# ======================================================
# CONSOLIDAR
# ======================================================
def consolidar(actividad):
    master = pd.read_parquet(RUTA_MASTER)

    for archivo in os.listdir("data"):
        if archivo.startswith("trabajo_") and actividad in archivo:
            ruta_archivo = os.path.join("data", archivo)
            df_temp = pd.read_parquet(ruta_archivo)

            columnas_necesarias = ["PK_ARTICULO", "ACTIVIDAD_COMERCIAL", "ACCION", "COMENTARIO"]
            for col in columnas_necesarias:
                if col not in df_temp.columns:
                    df_temp[col] = None

            df_temp = df_temp[columnas_necesarias].copy()

            master = master.merge(
                df_temp,
                on=["PK_ARTICULO", "ACTIVIDAD_COMERCIAL"],
                how="left",
                suffixes=("", "_NEW")
            )

            if "ACCION_NEW" in master.columns:
                master["ACCION"] = master["ACCION_NEW"].combine_first(master["ACCION"])

            if "COMENTARIO_NEW" in master.columns:
                master["COMENTARIO"] = master["COMENTARIO_NEW"].combine_first(master["COMENTARIO"])

            columnas_borrar = [c for c in ["ACCION_NEW", "COMENTARIO_NEW"] if c in master.columns]
            if columnas_borrar:
                master = master.drop(columns=columnas_borrar)

    master.to_parquet(RUTA_MASTER, index=False)


# ======================================================
# LOGIN
# ======================================================
if "login" not in st.session_state:
    st.session_state.login = False

if not st.session_state.login:
    user = st.text_input("Usuario")
    pwd = st.text_input("Password", type="password")

    if st.button("Ingresar"):
        with open("usuarios.json", encoding="utf-8") as f:
            usuarios = json.load(f)

        for u in usuarios:
            if u["usuario"] == user and u["password"] == pwd:
                st.session_state.login = True
                st.session_state.user = u
                st.rerun()

        st.error("Credenciales incorrectas")

else:
    actualizar_master()

    usuario = st.session_state.user["usuario"]
    rol = st.session_state.user["rol"]

    master = pd.read_parquet(RUTA_MASTER)

    st.success(f"Bienvenido {usuario}")

    # ==================================================
    # MASTER VACÍO
    # ==================================================
    if master.empty:
        if rol != "MASTER":
            st.warning("No existen Actividades Comerciales creadas")
            st.stop()

        st.markdown("### 🎯 Crear Primera Actividad")
        nueva_ac = st.text_input("Nombre Actividad")

        if st.button("CREAR ACTIVIDAD"):
            if not nueva_ac or not nueva_ac.strip():
                st.warning("Debe escribir un nombre de actividad")
                st.stop()

            if not os.path.exists(RUTA_BD):
                st.error("No existe BD_ACTUALIZACION.parquet")
                st.stop()

            bd = pd.read_parquet(RUTA_BD)

            bd["ACTIVIDAD_COMERCIAL"] = nueva_ac.strip()
            bd["MUNDO_AC"] = None
            bd["PRECIO_PROMOCIONAL"] = None
            bd["DESCUENTO"] = None
            bd["PORC_AHORRO"] = None
            bd["FECHA_INICIO"] = None
            bd["FECHA_FIN"] = None
            bd["ACCION"] = None
            bd["COMENTARIO"] = None

            estructura = pd.read_parquet(RUTA_MASTER).columns.tolist()

            for col in estructura:
                if col not in bd.columns:
                    bd[col] = None

            bd = bd[estructura]
            bd.to_parquet(RUTA_MASTER, index=False)

            st.success("Actividad creada")
            st.rerun()

    # ==================================================
    # SISTEMA NORMAL
    # ==================================================
    else:
        if rol == "MASTER":
            st.markdown("### 🎯 Crear Actividad")
            nueva_ac = st.text_input("Nueva Actividad")

            if st.button("CREAR"):
                if not nueva_ac or not nueva_ac.strip():
                    st.warning("Debe escribir un nombre de actividad")
                    st.stop()

                if nueva_ac.strip() in master["ACTIVIDAD_COMERCIAL"].dropna().unique():
                    st.error("Actividad ya existe")
                    st.stop()

                if not os.path.exists(RUTA_BD):
                    st.error("No existe BD_ACTUALIZACION.parquet")
                    st.stop()

                bd = pd.read_parquet(RUTA_BD)

                bd["ACTIVIDAD_COMERCIAL"] = nueva_ac.strip()
                bd["MUNDO_AC"] = None
                bd["PRECIO_PROMOCIONAL"] = None
                bd["DESCUENTO"] = None
                bd["PORC_AHORRO"] = None
                bd["FECHA_INICIO"] = None
                bd["FECHA_FIN"] = None
                bd["ACCION"] = None
                bd["COMENTARIO"] = None

                estructura = master.columns.tolist()
                for col in estructura:
                    if col not in bd.columns:
                        bd[col] = None

                bd = bd[estructura]

                master = pd.concat([master, bd], ignore_index=True)
                master.to_parquet(RUTA_MASTER, index=False)

                st.success("Actividad creada")
                st.rerun()

        actividades = master["ACTIVIDAD_COMERCIAL"].dropna().unique()

        if len(actividades) == 0:
            st.warning("No hay actividades disponibles")
            st.stop()

        actividad = st.selectbox("Seleccione Actividad", actividades)

        df = master[master["ACTIVIDAD_COMERCIAL"] == actividad].copy()

        if rol == "ADC":
            fam = st.session_state.user.get("familia", "ALL")
            if fam != "ALL":
                df = df[df["FAMILIA"] == fam]

        st.dataframe(df, use_container_width=True)

        # ==================================================
        # DESCARGAR EXCEL
        # ==================================================
        st.markdown("### 📥 Descargar Excel")

        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active

        for c, col in enumerate(df.columns, 1):
            ws.cell(row=1, column=c, value=col)

        for r, row in enumerate(df.itertuples(index=False), 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)

        wb.save(buffer)
        buffer.seek(0)

        st.download_button(
            "Descargar Excel",
            data=buffer,
            file_name=f"{actividad}_{usuario}.xlsx"
        )

        # ==================================================
        # SUBIR SOLO ADC Y JEFE_ADC
        # ==================================================
        if rol in ["ADC", "JEFE_ADC"]:
            st.markdown("### 📤 Subir archivo")
            st.warning("Después de subir el archivo, debe presionar CARGAR CAMBIOS")

            archivo = st.file_uploader("Subir Excel", type=["xlsx"])

            if archivo is not None:
                df_user = pd.read_excel(archivo)

                columnas_requeridas = ["PK_ARTICULO", "ACCION", "COMENTARIO"]
                if not all(c in df_user.columns for c in columnas_requeridas):
                    st.error("El archivo debe contener: PK_ARTICULO, ACCION y COMENTARIO")
                    st.stop()

                df_user["ACTIVIDAD_COMERCIAL"] = actividad

                ruta_salida = f"data/trabajo_{usuario}_{actividad}.parquet"
                df_user[["PK_ARTICULO", "ACTIVIDAD_COMERCIAL", "ACCION", "COMENTARIO"]].to_parquet(
                    ruta_salida,
                    index=False
                )

                st.success("Archivo cargado correctamente")

                if st.button("CARGAR CAMBIOS"):
                    consolidar(actividad)
                    st.success("Cambios aplicados")

        # ==================================================
        # DESCARGAS MASTER PARQUET
        # ==================================================
        if rol == "MASTER":
            st.markdown("### 📦 Descargar Base Parquet")

            col1, col2 = st.columns(2)

            with col1:
                with open(RUTA_MASTER, "rb") as f:
                    st.download_button(
                        "⬇️ Descargar MASTER TOTAL",
                        data=f,
                        file_name="MASTER_TOTAL.parquet"
                    )

            with col2:
                df_ac = master[master["ACTIVIDAD_COMERCIAL"] == actividad].copy()

                buffer_parquet = io.BytesIO()
                df_ac.to_parquet(buffer_parquet, index=False)
                buffer_parquet.seek(0)

                st.download_button(
                    "⬇️ Descargar Solo Actividad",
                    data=buffer_parquet,
                    file_name=f"MASTER_{actividad}.parquet"
                )
